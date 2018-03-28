# -*- coding: utf-8 -*-
#coding:utf-8
# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import codecs
import json
from openpyxl import Workbook

class JianzhiPipeline(object):
    def __init__(self):
        # 创建一个只写文件，指定文本编码格式为utf-8
        self.workbook = Workbook()
        self.bootsheet = self.workbook.active
        self.bootsheet.cell(1,1).value="标题"
        self.bootsheet.cell(1,2).value="酬劳"
        self.bootsheet.cell(1,3).value="内容"
        self.bootsheet.cell(1,4).value="工作地址"
        self.bootsheet.cell(1, 5).value = "详情查看及报名方式"

    def process_item(self, item, spider):

        title=item["title"]
        str_title = str(title).replace('u\'', '\'').replace("['",'').replace("']",'').replace("'" ,'').replace(',','').replace("'",'').replace("\n",'')
        str_title = str_title.decode("unicode-escape")




        if item['platform']=='58':
            address = item["address"]
            content = item["content"]
            temp_list=[str_title,item["money"],content,address,item["detailUrl"]]

        if item['platform']=='doumi':
            address = item["address"]
            str_address = str(address).replace('u\'', '\'').replace("['", '').replace("']", '').replace('(',
                                                                                                        '').replace(')',
                                                                                                                    '').replace(
                "'", '').replace(',', '').replace("'", '').replace("\n", '')
            str_address = str_address.decode("unicode-escape")
            content = item["content"]
            str_content = str(content).replace('u\'', '\'').replace("['", '').replace("']", '')
            str_content = str_content.decode("unicode-escape")

            temp_list=[str_title,item["money"],str_content,str_address,item["detailUrl"]]

        if item['platform']=='jianzhimao':
            address = item["address"]
            content = item["content"]
            str_content = str(content).replace('u\'', '\'').replace("['", '').replace("']", '')
            str_content = str_content.decode("unicode-escape")
            temp_list=[str_title,item["money"],str_content,address,item["detailUrl"]]
        if item['platform']=='jianke':
            address = item["address"]
            str_address = str(address).replace('u\'', '\'').replace("['", '').replace("']", '').replace('(',
                                                                                                        '').replace(')',
                                                                                                                    '').replace(
                "'", '').replace(',', '').replace("'", '').replace("\n", '')
            str_address = str_address.decode("unicode-escape")
            content = item["content"]
            str_content = str(content).replace('u\'', '\'').replace("['", '').replace("']", '')
            str_content = str_content.decode("unicode-escape")
            temp_list=[str_title,item["money"],str_content,str_address,item["detailUrl"]]
        if item['platform']=='jzba':
            address = item["address"]
            content = item["content"]
            temp_list=[str_title,item["money"],content,address,item["detailUrl"]]

        self.bootsheet.append(temp_list)
        self.workbook.save(item['platform']+"jianzhi.xlsx")
        return item

    def spider_closed(self, spider):
        self.workbook.close()