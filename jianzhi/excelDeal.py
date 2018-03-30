#coding:utf-8
from openpyxl import Workbook,load_workbook
import string
import datetime
class JianzhiDeal(object):
    def __init__(self):
        self.workbook = Workbook()
        self.bootsheet = self.workbook.active
        self.bootsheet.cell(1,1).value="标题"
        self.bootsheet.cell(1,2).value="酬劳"
        self.bootsheet.cell(1,3).value="内容"
        self.bootsheet.cell(1,4).value="工作地址"
        self.bootsheet.cell(1, 5).value = "详情查看及报名方式"
    def dealExcel(self,file):
        wb=load_workbook(file)
        sheet = wb.active
        nowTime = datetime.datetime.now().strftime('%Y-%m-%d')
        for row in sheet.rows:
            flag=0
            xrr=[]
            for cell in row:
                tempStr=cell.value
                xrr.append(cell.value)
                if tempStr is None:
                    continue
                if "模特" in tempStr.encode('UTF-8'):
                    flag=1
                    break
                if "代理" in tempStr.encode('UTF-8'):
                    flag=1
                    break
                if "试衣" in tempStr.encode('UTF-8'):
                    flag=1
                    break
                if "拍摄" in tempStr.encode('UTF-8'):
                    flag=1
                    break
                if "标题" in tempStr.encode('UTF-8'):
                    flag = 1
                    break
                if "拍" in tempStr.encode('UTF-8'):
                    flag=1
                    break
            if flag!=1:
                self.bootsheet.append(xrr)
        self.workbook.save(nowTime+"新鲜出炉的兼职信息.xlsx")
        wb.close()
        pass

    def deal58jianzhi(self,file):
        wb=load_workbook(file)
        sheet = wb.active
        nowTime = datetime.datetime.now().strftime('%Y-%m-%d')
        for row in sheet.rows:
            flag=1
            xrr=[]
            for cell in row:
                tempStr=cell.value
                xrr.append(cell.value)
                if tempStr is None:
                    continue
                if "发单" in tempStr.encode('UTF-8'):
                    flag=0
                    continue
                if "传单" in tempStr.encode('UTF-8'):
                    flag=0
                    continue
                if "客服" in tempStr.encode('UTF-8'):
                    flag = 0
                    continue
                if "家教" in tempStr.encode('UTF-8'):
                    flag = 0
                    continue
                if "老师" in tempStr.encode('UTF-8'):
                    flag = 0
                    continue
            if flag!=1:
                self.bootsheet.append(xrr)
        self.workbook.save(nowTime+"新鲜出炉的兼职信息.xlsx")
        wb.close()
        pass
if __name__ == '__main__':
    print "开始处理"
    dealClass=JianzhiDeal()
    #'58jianzhi.xlsx',
    filenames=['doumijianzhi.xlsx','jiankejianzhi.xlsx','jianzhimaojianzhi.xlsx','jzbajianzhi.xlsx']
    dealClass.deal58jianzhi('58jianzhi.xlsx')
    for file in filenames:
        dealClass.dealExcel(file)